#!/usr/bin/env python3
"""Build the deterministic 1,000-row expanded FETCH flip benchmark.

No model or network call is made. The first 200 rows are the prior flip
benchmark; 800 additional candidates are generated from deliberately reviewed
scenario families grounded in the four-label human-review workbook.
"""

from __future__ import annotations

import argparse
import csv
import json
import random
from collections import Counter
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook


HERE = Path(__file__).resolve().parent
REPO = HERE.parent
DEFAULT_WORKBOOK = REPO / "silver_labels/10_four_label_human_review/redaction_reviewed_v5_clean_four_label_human_review.xlsx"
DEFAULT_LEGACY = Path("/home/quinten/fetch/promptfoo/individual_facts/classification_flip_scenarios.csv")
DEFAULT_TAXONOMY = Path("/home/quinten/fetch/app/data/taxonomy_detailed_descriptions.csv")

OUT_CSV = HERE / "candidates/expanded_flip_candidates_1000.csv"
OUT_JSONL = HERE / "candidates/expanded_flip_candidates_1000.jsonl"
OUT_PROFILE = HERE / "analysis/source_dataset_profile.json"
OUT_MULTI = HERE / "analysis/multilabel_seed_rows.csv"

FIELDS = [
    "scenario_id", "opening_query_group", "opening_query", "initial_category",
    "initial_subcategory", "final_category", "final_subcategory", "hidden_fact",
    "fact_as_answer", "relevant_question_topic", "counterfactual_category",
    "counterfactual_subcategory", "counterfactual_hidden_fact",
    "counterfactual_fact_as_answer", "swap_pair", "direction", "source_stratum",
    "source_workbook_rows", "source_legacy_scenario_id", "multi_label_grounded",
    "paper_failure_mode", "domestic_violence_hidden_fact", "candidate_status",
]


@dataclass(frozen=True)
class Label:
    category: str
    subcategory: str


@dataclass(frozen=True)
class Family:
    name: str
    a: Label
    b: Label
    openings_a: tuple[str, ...]
    openings_b: tuple[str, ...]
    facts_a: tuple[str, ...]
    facts_b: tuple[str, ...]
    topic: str
    seed_rows: tuple[int, ...]
    multi: bool = False
    failure: str = ""
    dv_b: bool = False


def L(category: str, subcategory: str) -> Label:
    return Label(category, subcategory)


FAMILIES = (
    Family("divorce_vs_restraining", L("Family Law", "General (Divorce/Separation)"), L("Family Law", "Restraining Orders"),
           ("I need advice about separating from my partner after months of conflict.", "My spouse and I cannot keep living together and I need legal help.", "Things at home have become unbearable and I want to leave the relationship.", "I need to know my options concerning my marriage and our living situation."),
           ("I need help getting distance from my partner because I no longer feel safe.", "My ex keeps showing up and I need to know what legal protection is available.", "A relationship dispute has escalated and I need help keeping the other person away.", "I am afraid of what my partner may do and need legal advice."),
           ("There has been no threat or abuse; I only want to end the marriage.", "I am asking about divorce papers and dividing our property, not a protection order.", "The dispute is about separation and support, and nobody has threatened me.", "I want to dissolve the marriage and work out the terms of separation."),
           ("My partner threatened to hurt me if I leave and has blocked the door before.", "My spouse shoved me last night and I am afraid to go home.", "My ex has threatened me repeatedly and keeps coming to my house.", "My partner controls my money and has threatened violence when I ask to leave."),
           "whether there are threats, violence, coercive control, or an immediate need for protection", (51, 422), True, "domestic_violence_low_question_coverage", True),
    Family("custody_vs_restraining", L("Family Law", "Child Custody/Visitation"), L("Family Law", "Restraining Orders"),
           ("My child's other parent and I are fighting about where our child should stay.", "I need help changing the arrangements with my child's other parent.", "The other parent has been causing problems and I need advice about my child.", "I am worried about the current parenting arrangement and need legal help."),
           ("My child's other parent will not leave me alone and I need legal help.", "There is a serious dispute with my ex involving me and our child.", "I need help keeping my former partner away while we sort out issues involving our child.", "The situation with my child's other parent has escalated and I am scared."),
           ("There have been no threats; the only dispute is the parenting schedule.", "I need a custody order because the other parent will not return our child on time.", "The issue is which parent the child should live with and when visits occur.", "I want to modify custody and parenting time, not seek a protective order."),
           ("The other parent hit me while I was holding our child and threatened to come back.", "My ex threatened to take the child by force and hurt me.", "There is already violence at exchanges and I need immediate protection.", "My former partner has been stalking me and threatening our child's safety."),
           "whether the dispute concerns parenting time or violence and protection", (8, 38, 51), True, "domestic_violence_low_question_coverage", True),
    Family("separation_vs_financial_abuse", L("Family Law", "General (Divorce/Separation)"), L("Family Law", "Restraining Orders"),
           ("I want to separate, but I do not know how I can manage on my own.", "My partner handles all of our finances and I need advice about leaving.", "I need help ending my relationship and figuring out the next steps.", "I am thinking about divorce because our disagreements keep getting worse."),
           ("I need to leave my relationship but I am afraid I cannot do it safely.", "My spouse controls everything at home and I need legal advice.", "I want help getting away from my partner and protecting myself.", "My relationship has become controlling and I need to understand my options."),
           ("We have separate bank accounts and the disagreement is only about how to divide property.", "Nobody has restricted my money or threatened me; I just want a divorce.", "I can leave safely and need advice about support and property division.", "The only legal issue is ending the marriage and dividing our assets."),
           ("My partner cut off my access to every account and threatens me when I ask for money.", "I am not allowed to work or use our money, and my spouse threatens to hurt me if I leave.", "My partner takes my paychecks, tracks every purchase, and has threatened me.", "I have no access to money or identification because my spouse keeps them from me."),
           "whether the partner uses money, threats, or documents to control the caller", (51,), True, "domestic_violence_low_question_coverage", True),
    Family("stalking_vs_restraining", L("General Litigation", "Stalking Orders"), L("Family Law", "Restraining Orders"),
           ("Someone keeps contacting me and showing up near my home after I asked them to stop.", "I am being followed and watched by someone I know.", "A person keeps sending messages and appearing at places I go.", "I need legal protection from someone who will not leave me alone."),
           ("A person close to me keeps contacting and threatening me.", "Someone from my personal life is following me and I need protection.", "I am scared because a person I know keeps coming to my house.", "I need an order to keep someone away from me and my family."),
           ("The person is a neighbor I have never dated and is not related to me.", "This is a stranger who began following me after we met once in public.", "The person is a former coworker, not a family member or intimate partner.", "The person is an acquaintance with no family or dating relationship to me."),
           ("The person is my current spouse and we live together.", "This is my former dating partner, who threatened me during the relationship.", "The person is my child's other parent and has assaulted me before.", "The person is a family member who lived with me and has threatened violence."),
           "the caller's relationship to the person and any history of abuse", (70, 73, 387, 422), True, "criminal_vs_restraining_wrong_stalking_default", True),
    Family("criminal_case_vs_protection", L("Criminal Law", "Misdemeanor"), L("Family Law", "Restraining Orders"),
           ("There was a confrontation and now I need help with what happened.", "I got into an argument with someone and the police became involved.", "A dispute became physical and I do not know what legal help I need.", "I need advice after a serious incident with someone I know."),
           ("Someone I know has threatened me and I need legal help.", "A conflict has escalated and I am worried the person will come back.", "I need help after repeated unwanted contact and threats.", "I am afraid after an altercation and do not know what kind of order I need."),
           ("I was cited for misdemeanor assault and need a defense lawyer.", "The police arrested me and I am the person facing criminal charges.", "I have a court date on a criminal harassment charge.", "The prosecutor filed a misdemeanor case against me after the incident."),
           ("The person is my spouse, who threatened me, and I am seeking protection rather than a defense.", "My former partner assaulted me and keeps coming to my home.", "I am the person asking to be protected from a family member who threatened violence.", "My child's other parent is threatening me and I need an order keeping them away."),
           "whether the caller faces charges or seeks protection and the relationship involved", (418, 419, 422), True, "criminal_vs_restraining_wrong_stalking_default", True),
    Family("employment_vs_state_hearing", L("Labor & Employment", "Discrimination"), L("Administrative Law", "General (State)"),
           ("I was treated unfairly at work and filed a complaint about it.", "My employer discriminated against me and I need help with the next step.", "I reported unfair treatment at my job and now there is a legal process.", "I need an attorney for a workplace discrimination matter."),
           ("I have a hearing about a complaint involving my workplace.", "A government office sent me papers after I reported my employer.", "My employment complaint is moving forward and I need representation.", "I need help responding to a decision concerning a workplace complaint."),
           ("No agency hearing is pending; I need to bring a discrimination claim against my employer.", "My issue is directly with the employer and I have not appealed any state decision.", "I want advice about workplace discrimination before filing with an agency.", "The employer's discriminatory treatment is the only matter I need help with."),
           ("A state agency denied my complaint and scheduled an administrative appeal hearing.", "I am appealing the state bureau's final order, not suing my employer right now.", "The notice is for a contested-case hearing before a state administrative law judge.", "My immediate deadline is an appeal of the state agency's decision on my complaint."),
           "whether a state agency has issued a decision or scheduled an administrative hearing", (83,), True, "employment_admin_ignores_hearing_signals"),
    Family("wrongful_discharge_vs_unemployment", L("Labor & Employment", "Wrongful Discharge"), L("Administrative Law", "Unemployment"),
           ("I was fired and need help because I have no income now.", "My employer let me go and I think the reason was unfair.", "I lost my job after a dispute and need legal advice.", "I was terminated recently and do not know what rights I have."),
           ("I was fired and then received a denial letter about benefits.", "After losing my job, I applied for help and was turned down.", "I need assistance after my termination and a decision about my income.", "I have a deadline connected to losing my job and need a lawyer."),
           ("I am challenging the employer's retaliatory firing, not a benefits decision.", "I was fired for reporting misconduct and have not applied for unemployment.", "My claim is that the employer terminated me unlawfully because of discrimination.", "The only dispute is whether my employer was allowed to fire me."),
           ("The state denied my unemployment claim and I need to appeal that decision.", "My hearing is about unemployment benefits, not a lawsuit against the employer.", "The employment department says I am ineligible for benefits and set an appeal deadline.", "I need representation at a state unemployment-benefits hearing."),
           "whether the dispute is with the employer or the state unemployment agency", (12, 27, 32, 342), True, "employment_admin_ignores_hearing_signals"),
    Family("work_injury_vs_personal_injury", L("Workers' Comp", "State"), L("General Litigation", "Personal Injury"),
           ("I was hurt while doing a job and need to know who should pay my medical bills.", "I injured myself at a work site and have not been able to return.", "I had an accident while helping at a business and need legal advice.", "I got badly hurt during a shift and do not know what claim to make."),
           ("I was injured at a store and the business says it is not responsible.", "I got hurt on someone else's property while I was there for work.", "An accident left me with medical bills and I need help recovering damages.", "I was injured while completing a task and need to know whether I can sue."),
           ("I was an employee, on the clock, and performing my assigned duties when it happened.", "The injury occurred during my scheduled shift for my employer.", "I receive wages from the company and was doing exactly what my supervisor asked.", "I was clocked in at my regular workplace when the accident occurred."),
           ("I was an independent contractor and was not the company's employee.", "I was off the clock on a personal errand when the accident happened.", "I was a customer visiting the property, not a worker there.", "A separate company's careless driver injured me, and my employer did not cause it."),
           "employment status, whether the caller was on duty, and who caused the injury", (28, 349, 420), True, "injury_location_does_not_flip_off_clock"),
    Family("workers_comp_vs_third_party", L("Workers' Comp", "State"), L("Workers' Comp", "Third Party litigation"),
           ("I was injured during my shift and workers' compensation is involved.", "I got hurt at work and need to know whether I have another claim.", "A workplace accident caused serious injuries and I need legal advice.", "I have a workers' compensation case after being hurt on the job."),
           ("I was injured at work because of equipment or a person from outside my company.", "My work injury involved another business and I need to know who is responsible.", "I was hurt on the job and someone says a separate lawsuit may be possible.", "An accident during work involved people who do not work for my employer."),
           ("No outside person or company caused it; I slipped while doing my ordinary job.", "The accident involved only my employer and normal workplace conditions.", "There is no defective product or third party involved in the injury.", "I was hurt by a condition controlled entirely by my employer."),
           ("A delivery driver from another company caused the accident while I was working.", "Defective equipment made by an outside manufacturer caused my injury.", "A subcontractor who is not my employer created the dangerous condition.", "The customer who attacked me was not employed by my company."),
           "whether someone other than the employer caused the work injury", (28, 349, 420), True, "injury_location_does_not_flip_off_clock"),
    Family("bankruptcy_vs_creditor", L("Bankruptcy", "Personal"), L("Debtor/Creditor", "General (Creditor)"),
           ("Money problems are getting worse and I need legal advice about what to do.", "I am overwhelmed by unpaid debts and cannot keep up financially.", "There are several bills involved and I do not know what legal option fits.", "I need help with a serious debt problem that is affecting my finances."),
           ("Someone has not paid money that is owed and I need legal help.", "A debt has gone unpaid and I want to understand my options.", "I am having trouble getting payment and the situation is hurting me financially.", "I need advice about money that remains unpaid."),
           ("I am the person who owes several creditors and want to file bankruptcy.", "The debts are mine, and I cannot afford the payments or collection lawsuits.", "I need relief from my own credit cards and medical bills.", "Creditors are suing me and I am considering personal bankruptcy."),
           ("I am the person owed money and want to collect from the debtor.", "The unpaid invoices are owed to me for work I completed.", "I loaned the other person money and they have stopped repaying me.", "My customer owes my business money and I want to pursue collection."),
           "whether the caller owes the debt or is trying to collect it", (108, 110, 120), True, "bankruptcy_collections_role_confusion"),
    Family("judgment_vs_bankruptcy", L("Debtor/Creditor", "Judgement Collection"), L("Bankruptcy", "Personal"),
           ("A court case about money ended, but the financial problem is not resolved.", "There is already a decision about a debt and I need to know the next step.", "I have court papers involving money and cannot figure out what to do now.", "A judgment has created a serious financial issue for me."),
           ("I am facing a court judgment and several other debts I cannot handle.", "A money judgment is part of a larger financial crisis for me.", "I need advice about a judgment and whether there is a broader solution.", "Court-ordered debt is causing financial problems and I need help."),
           ("The judgment is in my favor and I need to enforce it against the debtor.", "I won the case, but the other person still has not paid me.", "The judge ordered payment to me and I need help collecting it.", "I am the judgment creditor and want to garnish or locate assets."),
           ("The judgment is against me and I have many other debts, so I want to file bankruptcy.", "I am the debtor, cannot pay the judgment, and need bankruptcy advice.", "Several creditors have judgments against me and I cannot meet basic expenses.", "I am considering personal bankruptcy because I owe this judgment and other bills."),
           "who won the judgment and whether the caller needs debt relief", (108, 110), True, "bankruptcy_collections_role_confusion"),
    Family("tenant_vs_landlord", L("Real Property", "Tenant (Residential)"), L("Real Property", "Landlord Residential"),
           ("There is a dispute about a rental home and repairs have not been made.", "I need help with a lease problem involving the place where someone lives.", "The other person is not following the rental agreement and I need advice.", "A housing dispute has reached the point where I may need a lawyer."),
           ("I am responsible for a rental property and there is a serious lease dispute.", "Someone in a rental unit is causing problems and I need legal help.", "I need to enforce terms of a residential rental agreement.", "There is a dispute over possession of a rented home."),
           ("I pay rent to the owner and live in the unit as the tenant.", "I do not own the property; I signed the lease as the renter.", "The landlord is the other party and I am asking for repairs to my home.", "I am the tenant facing eviction from the apartment where I live."),
           ("I own the property and rent the unit to the other person.", "I am the landlord and the tenant has stopped paying rent.", "The lease lists me as the property owner renting the home out.", "I manage the rental for its owner and need to address the tenant's breach."),
           "whether the caller rents the home or owns and rents it to someone else", (14, 365, 371, 419), True, ""),
    Family("dui_vs_dmv", L("Criminal Law", "DUII/DWS"), L("Administrative Law", "DMV (Non-Criminal)"),
           ("My driving privileges are in danger and I have a deadline coming up.", "I received papers about my license and need legal advice.", "I cannot drive right now and need help challenging the decision.", "There is a hearing about whether I can keep my driver's license."),
           ("The DMV says there is a problem with my license and I need to appeal.", "A state office refused to renew my driving privileges.", "I need help with a noncriminal decision affecting my license.", "My license was suspended and I do not understand which process applies."),
           ("I was arrested for driving under the influence and face a criminal charge.", "The suspension followed a DUII arrest and I have a criminal court date.", "Police cited me for DUII after a traffic stop.", "This involves criminal charges for impaired driving, not only DMV paperwork."),
           ("There was no arrest or DUII; the DMV acted because of a medical review.", "The issue is rejected identity documents at the DMV, with no criminal case.", "I am appealing an age-based license renewal denial and was not charged with a crime.", "This is only a DMV administrative hearing about driving fitness."),
           "whether there is a DUII arrest or only a noncriminal DMV decision", (422,), False, ""),
    Family("custody_vs_support", L("Family Law", "Child Custody/Visitation"), L("Family Law", "Child Support/Modification"),
           ("My child's other parent and I disagree about our current arrangement.", "I need to change an order involving my child.", "There is an ongoing dispute with the other parent and I need legal help.", "The current parenting arrangement is not working and I need advice."),
           ("I need help enforcing part of an order involving my child.", "My child's other parent is not following what the court ordered.", "There is a problem with the agreement we have concerning our child.", "I need to modify the terms of a family court order."),
           ("The dispute is about where the child lives and when each parent has visits.", "I need more parenting time, and money is not the issue.", "The other parent is refusing visits required by the custody order.", "I am asking the court to change custody because the child moved."),
           ("The only dispute is the amount of monthly child support.", "Parenting time is settled, but support payments have stopped.", "I need to modify child support after my income changed.", "The other parent owes past-due support, and custody is not disputed."),
           "whether the issue is parenting time or financial support", (8, 51), True, ""),
    Family("wages_vs_wrongful_discharge", L("Labor & Employment", "Wage and Hour Claims"), L("Labor & Employment", "Wrongful Discharge"),
           ("I had a problem with my pay and raised it with my manager.", "My employer and I are in a dispute after I questioned my paycheck.", "I need legal advice about something that happened after I complained at work.", "There is a serious issue involving my hours and my employer's response."),
           ("I complained about work and then my employer took action against me.", "A workplace dispute about money escalated and I need a lawyer.", "My manager reacted badly when I raised a problem at work.", "I need help after reporting an issue with my employment."),
           ("I still have the job; the issue is unpaid overtime and deleted hours.", "I am only trying to recover wages the employer failed to pay.", "My paycheck omitted hours I worked, but I was not terminated.", "The employer misclassified me and has not paid required overtime."),
           ("The employer fired me immediately after I complained about the missing wages.", "I was terminated in retaliation for reporting illegal pay practices.", "My manager discharged me after I asked the labor agency about my pay.", "The unpaid wages matter is resolved, but I need help with the retaliatory firing."),
           "whether employment continues and whether the caller was fired for complaining", (12, 32, 344), True, ""),
    Family("malpractice_vs_wrongful_death", L("General Litigation", "Malpractice-Medical"), L("General Litigation", "Wrongful Death"),
           ("A hospital made a serious mistake during treatment and my family needs help.", "I believe negligent medical care caused terrible harm to a relative.", "Something went badly wrong at the hospital and I want legal advice.", "A doctor failed to provide proper care and the consequences were severe."),
           ("My family is considering a claim after a medical emergency.", "A relative suffered catastrophic harm following hospital care.", "We need an attorney to review what happened during medical treatment.", "The hospital's actions caused a devastating outcome for our family."),
           ("The patient survived and needs compensation for additional injuries caused by the treatment.", "My relative is alive but now has permanent injuries from the medical error.", "The negligent procedure caused more treatment and disability, but no one died.", "I am the injured patient and want to bring a medical malpractice claim."),
           ("The patient died from the injuries, and the estate wants to bring a claim.", "My relative passed away after the medical error and we seek damages for the death.", "The negligent care caused death, and I am contacting counsel for the family.", "The patient never recovered and died; we need a wrongful-death attorney."),
           "whether the patient survived and who would bring the claim", (24,), True, ""),
    Family("paternity_vs_custody", L("Family Law", "Paternity"), L("Family Law", "Child Custody/Visitation"),
           ("I need help establishing my legal rights concerning a young child.", "I am not listed in some of my child's records and need legal advice.", "There is a dispute about my role as a parent and what I can do.", "I need a family lawyer to help secure my rights involving my child."),
           ("The other parent is preventing me from seeing my child.", "I need help getting a court order involving my child.", "My child's living arrangement has changed and I need legal advice.", "I want to establish a regular arrangement with my child's other parent."),
           ("My main issue is proving I am the legal parent and getting onto the birth certificate.", "Parentage has never been legally established, and that must happen first.", "I need genetic testing and a paternity judgment.", "Nobody disputes visits yet; I need the court to determine legal parentage."),
           ("Paternity is already established; the dispute is over custody and visits.", "I am on the birth certificate and only need a parenting-time order.", "Both parties agree I am the parent, but the other parent is withholding visits.", "Legal parentage is settled and I need to modify custody."),
           "whether parentage is legally established and whether custody is disputed", (23, 38, 51), True, ""),
    Family("guardianship_vs_elder_abuse", L("Wills & Trusts", "Conservatorship/Guardianship"), L("Wills & Trusts", "Elder Abuse"),
           ("An older relative can no longer manage important decisions and the family needs help.", "There is a dispute about who should handle an elderly parent's affairs.", "My parent has dementia and I do not know who has authority to act.", "I need legal advice about protecting a relative who cannot care for themself."),
           ("Someone is controlling an elderly relative's finances and I am concerned.", "A family member has taken over my parent's affairs and will not explain anything.", "I think an older relative is being taken advantage of and need legal help.", "There are suspicious changes involving a vulnerable parent's money and decision-making."),
           ("No money is missing; we need a court-appointed guardian because there is no valid power of attorney.", "The issue is legal authority to arrange care for a parent with dementia.", "We need a conservatorship so someone can manage property and medical decisions.", "The family agrees protection is needed, and the question is who should be guardian."),
           ("The person with authority has transferred my parent's money into their own account.", "A caregiver pressured my parent to change beneficiaries and took large payments.", "My sibling is using the power of attorney for personal purchases.", "A supposed friend isolated my parent and is selling property for their own benefit."),
           "whether the need is authority to act or suspected exploitation", (31, 71, 399, 404, 411, 421), True, ""),
    Family("ssd_vs_ssi", L("Administrative Law", "SSD (Social Security Disability)"), L("Administrative Law", "Social Security/SSI"),
           ("Social Security denied my disability benefits even though I cannot work.", "I need help appealing a disability decision from Social Security.", "My medical records show I am disabled, but my claim was denied.", "I received a denial for disability payments and have an appeal deadline."),
           ("A disability benefit based on my finances was denied and I need help.", "Social Security turned down my application and I do not understand why.", "I cannot work and have very little income, but my benefits were denied.", "I need an attorney for an appeal involving disability assistance."),
           ("I applied for SSDI based on my work history and payroll contributions.", "The denial concerns Social Security Disability Insurance from my prior employment.", "I have enough work credits and am appealing an SSD decision.", "This is SSDI, not the needs-based SSI program."),
           ("I applied for SSI because I have almost no income or resources.", "The denied benefit is Supplemental Security Income, not SSDI.", "I lack enough work credits and filed for the needs-based SSI program.", "My appeal concerns SSI eligibility based on disability and limited income."),
           "whether the benefit is work-credit SSDI or needs-based SSI", (94, 96, 397), True, ""),
    Family("copyright_vs_patent", L("Intellectual Property", "Trademark/Copyright"), L("Intellectual Property", "Patent (Reg Patent Attys Only)"),
           ("I created something new and want to protect it before showing it publicly.", "I need legal help protecting a project I have been developing.", "Someone may copy my work and I want to understand my rights.", "I want to register protection for something my business created."),
           ("I developed a new product and need advice about protecting the idea.", "My design may have commercial value and I need an intellectual property lawyer.", "I want to stop competitors from copying a new creation.", "I need help filing the right kind of protection for my work."),
           ("The work is a novel and its title and branding; there is no new machine or process.", "I need copyright protection for original artwork and a trademark for the brand.", "The creation is software code and written content that I want to register.", "I am protecting a book, logo, and licensing rights rather than an invention."),
           ("The creation is a new mechanical device and I need a registered patent attorney.", "I invented a novel manufacturing process and want to file a utility patent.", "The product has a new functional mechanism, not merely artwork or branding.", "I need a patent application for a new machine before publicly disclosing it."),
           "whether the creation is expressive branding/content or a functional invention", (10, 40, 282), True, ""),
    Family("contractor_vs_small_claims", L("Real Property", "Construction/Contractors"), L("Consumer Law", "Small Claims Advice"),
           ("A contractor did poor work on my home and we disagree about payment.", "My renovation is unfinished and I need to know how to recover the cost.", "There is a dispute over defective work by a home contractor.", "I paid for repairs that were not completed properly and need legal advice."),
           ("A small home-repair dispute may need to go to court and I need guidance.", "I want to recover money for bad contractor work but do not know which court to use.", "The amount at issue in a renovation dispute is limited and I need advice.", "A contractor owes me money for repairs and I may need to file a claim myself."),
           ("The dispute involves a construction lien and contractor licensing issues beyond small claims.", "I need counsel for a complex construction contract and damage exceeding the small-claims limit.", "The contractor's bond, licensing, and multiple subcontractors are involved.", "I need representation in circuit court over a large construction defect claim."),
           ("The only claim is for a small amount within the small-claims limit, and I want filing guidance.", "I plan to represent myself in small claims to recover the limited repair cost.", "The dispute is under the small-claims cap and no lien or licensing issue exists.", "I need advice on completing small-claims forms for this modest contractor dispute."),
           "amount, desired court, liens, licensing, and complexity", (59, 165, 368, 391), True, ""),
    Family("license_vs_employment_discrimination", L("Administrative Law", "Professional Licensing"), L("Labor & Employment", "Discrimination"),
           ("A problem with my professional work status is preventing me from earning a living.", "I need legal help after being told I cannot continue in my profession.", "A decision connected to my occupation has cost me work.", "I believe unfair treatment is keeping me from working in my field."),
           ("I was treated unfairly in connection with my job and qualifications.", "Bias may be the reason I cannot work in my profession now.", "I need help challenging discrimination that affected my ability to work.", "My career has been harmed by an unfair decision and I need a lawyer."),
           ("The state licensing board revoked my occupational license and I must appeal its order.", "My employer did not fire me; a government board denied my license renewal.", "The immediate issue is a character-and-fitness review by the professional licensing agency.", "I need to challenge the state board's suspension of my professional license."),
           ("My license remains valid, but my employer fired me because of a protected characteristic.", "The discriminatory decision was made by my employer, not the licensing board.", "I was denied a promotion because of my disability even though my license is active.", "My employer used age and race in hiring; no government license is in dispute."),
           "whether the decision-maker is a licensing board or an employer", (2, 83, 424), True, "employment_admin_ignores_hearing_signals"),
    Family("mold_vs_tenant", L("General Litigation", "Mold Claims"), L("Real Property", "Tenant (Residential)"),
           ("There are dangerous conditions in my rental home and the owner will not fix them.", "Problems in my apartment are making my family sick and I need legal help.", "My landlord has ignored repeated reports about conditions in the unit.", "The place I rent has serious water damage and I do not know what claim to bring."),
           ("My landlord is threatening action after I complained about conditions in my home.", "I need help enforcing my rights concerning repairs in my rental.", "A repair dispute with my landlord has escalated and I need advice.", "The owner will not maintain the apartment and says I am responsible."),
           ("Testing found toxic mold, and the main claim is for illness and remediation caused by mold exposure.", "The water damage produced confirmed mold that caused documented medical problems.", "I need damages specifically for toxic mold contamination and related injuries.", "An inspector identified extensive mold, and that is the focus of the lawsuit."),
           ("There is no mold; the dispute is the landlord's failure to repair plumbing and an eviction threat.", "The issue is habitability and rent withholding, with no mold-related injury claim.", "I need repairs and protection from retaliation as a residential tenant.", "The landlord served an eviction notice after I requested ordinary repairs."),
           "whether mold exposure is the claim or the issue is ordinary tenant rights", (14, 365, 371), True, ""),
    Family("defamation_vs_online_harassment", L("General Litigation", "Libel/Slander/Defamation"), L("General Litigation", "Online Harassment/Doxing/Bullying"),
           ("Someone is posting harmful things about me online and I need them to stop.", "False and upsetting material about me is spreading on social media.", "An online campaign against me is damaging my life and reputation.", "I need legal help with repeated posts and messages targeting me."),
           ("I am being targeted online and the conduct keeps escalating.", "Someone is using the internet to attack me and involve other people.", "Repeated online activity is making me fear for my safety.", "A person will not stop contacting and posting about me online."),
           ("The central issue is a specific false statement of fact that cost me clients.", "They published a knowingly false accusation, and I seek damages to my reputation.", "The post falsely says I committed a crime and employers have seen it.", "I need a defamation claim for false factual statements, not unwanted messages."),
           ("They published my home address and are directing strangers to contact me.", "The conduct is repeated threats and doxing rather than one reputational statement.", "They created many accounts to harass me and send private information to others.", "The main issue is coordinated cyberbullying and unwanted contact, not whether a statement is false."),
           "whether the harm is a false factual statement or repeated harassment and doxing", (73, 290), True, ""),
    Family("veterans_benefits_vs_va_loan", L("Administrative Law", "Military/Veterans"), L("Real Property", "Government Loans (VA,FHA,Etc.)"),
           ("I am a veteran having trouble with a government program and need legal advice.", "A VA-related decision is affecting my finances and I need help.", "I received a denial involving a veterans program and do not know how to appeal.", "I need an attorney familiar with veterans issues and a government decision."),
           ("There is a problem involving the VA and my home financing.", "A veterans program connected to my property has created a dispute.", "I need help with a VA decision that affects my housing.", "My status as a veteran is tied to a legal issue involving my home."),
           ("The VA stopped my disability benefits, and the dispute is about benefits eligibility.", "I am appealing denial of VA medical care and pension benefits.", "The issue is a Department of Veterans Affairs benefit termination, not a mortgage.", "I need restoration of veterans compensation that the VA discontinued."),
           ("The dispute concerns the terms and servicing of my VA-backed construction loan.", "My benefits are intact; the issue is approval of a VA home loan.", "I need advice about a lender's handling of a government-backed VA mortgage.", "The legal problem is a VA loan on residential property, not veterans compensation."),
           "whether the VA issue is a personal benefit or a government-backed property loan", (384, 424), True, ""),
    Family("medicaid_vs_guardianship", L("Administrative Law", "Medicare/Medicaid"), L("Wills & Trusts", "Conservatorship/Guardianship"),
           ("My elderly parent may lose care and our family does not know what to do.", "A care facility says there is a legal problem involving my parent's property.", "My parent has dementia and a benefits problem threatens the current placement.", "We need urgent advice so an older relative can continue receiving care."),
           ("No one currently has authority to make decisions for my parent with dementia.", "Our family cannot act for an incapacitated parent and care decisions are overdue.", "An older relative cannot understand the consequences of an important decision.", "We need legal authority to protect a parent who can no longer manage affairs."),
           ("The immediate issue is a Medicaid termination notice based on assets, and an appeal is due.", "A government health program denied continued coverage for the care facility.", "We already have valid authority to act, but Medicaid says benefits will end.", "The dispute is whether the state may count the house for Medicaid eligibility."),
           ("Medicaid has not denied anything; the problem is that nobody can legally consent or manage property.", "We need a guardian because there is no power of attorney and our parent lacks capacity.", "The court must appoint someone to make care and financial decisions.", "The only issue is obtaining a conservatorship for the incapacitated adult."),
           "whether benefits were denied or the family lacks legal authority to act", (404, 411), True, ""),
    Family("business_bankruptcy_vs_dissolution", L("Bankruptcy", "Business"), L("Business & Corporate", "Sale of Business"),
           ("My small business is closing with debts and I need advice about winding it down.", "A former partner and I need to resolve ownership and unpaid business obligations.", "The company cannot continue operating and we do not know the right legal process.", "I need help ending a business after it ran into serious financial trouble."),
           ("I am leaving a company and need to transfer my ownership interest properly.", "My business partner and I agreed to separate but the paperwork was never completed.", "We need to close or transfer a small business and settle our agreement.", "There is a dispute over buying out an owner's share of a failing company."),
           ("The company cannot pay creditors and needs to file a business bankruptcy case.", "The main issue is insolvency and bankruptcy relief for the business debts.", "Creditors are suing the company, which has no assets left to pay them.", "We need counsel for a business bankruptcy, not merely ownership-transfer documents."),
           ("The business can pay its debts; I need documents transferring my ownership to the other partner.", "The issue is the buyout and sale of my company interest, not insolvency.", "We agreed on a purchase price and need to complete the ownership transfer.", "No bankruptcy is planned; I need help selling and dissolving the business properly."),
           "whether the company is insolvent or the issue is ownership transfer and dissolution", (33, 109), True, "bankruptcy_collections_role_confusion"),
)


CLOSINGS = ("", " What are my options?", " I do not know what to do next.", " Can a lawyer help me?", " I need to understand my rights.")

LEGACY_LABEL_MAP = {
    ("Bankruptcy", "General (Bankruptcy)"): ("Bankruptcy", "Personal"),
    ("Business and Corporate", "General (contracts, entities)"): ("Business & Corporate", "General (Contracts, Business, Organization)"),
    ("Consumer Law", "General Consumer"): ("Consumer Law", "General"),
    ("Criminal Law", "Assault/Battery (Non-DV)"): ("Criminal Law", "Misdemeanor"),
    ("Criminal Law", "DUI"): ("Criminal Law", "DUII/DWS"),
    ("Debtor/Creditor", "Judgment Collection"): ("Debtor/Creditor", "Judgement Collection"),
    ("Family Law", "Domestic Violence"): ("Family Law", "Restraining Orders"),
    ("Family Law", "General (Divorce/Separation etc.)"): ("Family Law", "General (Divorce/Separation)"),
    ("Labor & Employment", "Discrimination - Employee"): ("Labor & Employment", "Discrimination"),
    ("Labor & Employment", "Wrongful Discharge - Employee"): ("Labor & Employment", "Wrongful Discharge"),
    ("Real Property", "Landlord (Residential)"): ("Real Property", "Landlord Residential"),
    ("Workers' Comp", "General (Workers' Comp)"): ("Workers' Comp", "State"),
}


def normalize_label(category: str, subcategory: str) -> tuple[str, str]:
    return LEGACY_LABEL_MAP.get((category, subcategory), (category, subcategory))


def workbook_profile(path: Path) -> tuple[dict, list[dict]]:
    ws = load_workbook(path, read_only=True, data_only=True).active
    values = ws.iter_rows(values_only=True)
    header = next(values)
    ix = {name: pos for pos, name in enumerate(header)}
    multi_rows, texts = [], []
    label_count = Counter()
    for row_number, row in enumerate(values, 2):
        text = str(row[ix["problem_description"]] or "").strip()
        texts.append(text)
        labels = []
        for slot in range(1, 5):
            c = row[ix[f"four_label_ai_category_{slot}"]]
            s = row[ix[f"four_label_ai_subcategory_{slot}"]]
            if c and s:
                labels.append(f"{c} > {s}")
        label_count[len(labels)] += 1
        if len(labels) >= 2:
            multi_rows.append({
                "workbook_row": row_number,
                "problem_description": text,
                "candidate_label_count": len(labels),
                "candidate_labels": " | ".join(labels),
                "review_priority": row[ix["internal_review_priority"]] or "",
            })
    word_counts = sorted(len(t.split()) for t in texts)
    profile = {
        "workbook": str(path.relative_to(REPO)),
        "data_rows": len(texts),
        "unique_problem_descriptions": len(set(texts)),
        "problem_description_word_count": {
            "min": word_counts[0], "median": word_counts[len(word_counts) // 2],
            "mean": round(sum(word_counts) / len(word_counts), 2), "max": word_counts[-1],
        },
        "four_label_candidate_count_distribution": dict(sorted(label_count.items())),
        "rows_with_two_or_more_candidate_labels": len(multi_rows),
        "benchmark_seed_rows": sorted({r for family in FAMILIES for r in family.seed_rows}),
    }
    return profile, multi_rows


def write_rows(path: Path, rows: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=FIELDS)
        writer.writeheader()
        writer.writerows({field: row.get(field, "") for field in FIELDS} for row in rows)


def expanded_rows() -> list[dict]:
    rows = []
    for family_index, family in enumerate(FAMILIES, 1):
        # 22 families x 30 rows + 5 families x 28 rows = 800. Both
        # directions remain exactly balanced within every family.
        per_direction = 15 if family_index <= 22 else 14
        for direction, initial, final, openings, final_facts, counter_facts in (
            ("AtoB", family.a, family.b, family.openings_a, family.facts_b, family.facts_a),
            ("BtoA", family.b, family.a, family.openings_b, family.facts_a, family.facts_b),
        ):
            for i in range(per_direction):
                opening = openings[i % len(openings)] + CLOSINGS[i // len(openings)]
                fact = final_facts[(i * 3 + family_index) % len(final_facts)]
                counter = counter_facts[(i * 5 + family_index) % len(counter_facts)]
                seq = len(rows) + 1
                dv = family.dv_b and final == family.b
                rows.append({
                    "scenario_id": f"x{seq:04d}",
                    "opening_query_group": f"f{family_index:02d}_{direction}_{i+1:02d}",
                    "opening_query": opening,
                    "initial_category": initial.category,
                    "initial_subcategory": initial.subcategory,
                    "final_category": final.category,
                    "final_subcategory": final.subcategory,
                    "hidden_fact": fact,
                    "fact_as_answer": fact,
                    "relevant_question_topic": family.topic,
                    "counterfactual_category": initial.category,
                    "counterfactual_subcategory": initial.subcategory,
                    "counterfactual_hidden_fact": counter,
                    "counterfactual_fact_as_answer": counter,
                    "swap_pair": family.name,
                    "direction": direction,
                    "source_stratum": "expanded_workbook_grounded",
                    "source_workbook_rows": ";".join(map(str, family.seed_rows)),
                    "source_legacy_scenario_id": "",
                    "multi_label_grounded": str(family.multi).lower(),
                    "paper_failure_mode": family.failure,
                    "domestic_violence_hidden_fact": str(dv).lower(),
                    "candidate_status": "synthetic_candidate_needs_human_validation",
                })
    assert len(rows) == 800, len(rows)
    return rows


def legacy_rows(path: Path) -> list[dict]:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        source = list(csv.DictReader(handle))
    if len(source) != 200:
        raise ValueError(f"Expected 200 legacy rows, found {len(source)}")
    rows = []
    hard = {"criminal_vs_restraining", "employment_admin", "injury_location", "bankruptcy_vs_collections", "domestic_violence"}
    for i, old in enumerate(source, 1):
        initial = normalize_label(old["initial_category"], old["initial_subcategory"])
        final = normalize_label(old["final_category"], old["final_subcategory"])
        dv = old["swap_pair"] == "domestic_violence" and old["direction"] == "AtoB"
        rows.append({
            "scenario_id": f"legacy_{i:03d}",
            "opening_query_group": f"legacy_{old['scenario_id']}",
            "opening_query": old["opening_query"],
            "initial_category": initial[0], "initial_subcategory": initial[1],
            "final_category": final[0], "final_subcategory": final[1],
            "hidden_fact": old["hidden_fact"], "fact_as_answer": old["fact_as_answer"],
            "relevant_question_topic": old["relevant_question_topic"],
            "counterfactual_category": initial[0], "counterfactual_subcategory": initial[1],
            "counterfactual_hidden_fact": f"The clarifying answer confirms the original {initial[0]} issue and does not add the fact that would support {final[0]}.",
            "counterfactual_fact_as_answer": f"No. My issue is the original {initial[0]} matter, not the alternative described.",
            "swap_pair": old["swap_pair"], "direction": old["direction"],
            "source_stratum": "legacy_200_snapshot", "source_workbook_rows": "",
            "source_legacy_scenario_id": old["scenario_id"], "multi_label_grounded": "false",
            "paper_failure_mode": old["swap_pair"] if old["swap_pair"] in hard else "",
            "domestic_violence_hidden_fact": str(dv).lower(),
            "candidate_status": "legacy_candidate_needs_revalidation",
        })
    return rows


def validate(rows: list[dict], taxonomy_path: Path) -> dict:
    with taxonomy_path.open(newline="", encoding="utf-8-sig") as handle:
        taxonomy = {(r["Category"], r["Subcategory"]) for r in csv.DictReader(handle)}
    errors = []
    ids = Counter(row["scenario_id"] for row in rows)
    if len(rows) != 1000:
        errors.append(f"expected 1000 rows; found {len(rows)}")
    if any(n != 1 for n in ids.values()):
        errors.append("scenario IDs are not unique")
    for row in rows:
        for prefix in ("initial", "final", "counterfactual"):
            label = (row[f"{prefix}_category"], row[f"{prefix}_subcategory"])
            if label not in taxonomy:
                errors.append(f"{row['scenario_id']}: invalid {prefix} label {label}")
        if (row["initial_category"], row["initial_subcategory"]) == (row["final_category"], row["final_subcategory"]):
            errors.append(f"{row['scenario_id']}: initial and final labels are identical")
        if (row["initial_category"], row["initial_subcategory"]) != (row["counterfactual_category"], row["counterfactual_subcategory"]):
            errors.append(f"{row['scenario_id']}: counterfactual does not oppose final outcome")
        if not all(str(row.get(field, "")).strip() for field in ("opening_query", "hidden_fact", "fact_as_answer", "relevant_question_topic", "counterfactual_hidden_fact")):
            errors.append(f"{row['scenario_id']}: missing required content")
    summary = {
        "row_count": len(rows),
        "source_strata": dict(Counter(r["source_stratum"] for r in rows)),
        "swap_pair_count": len(set(r["swap_pair"] for r in rows)),
        "swap_pair_distribution": dict(sorted(Counter(r["swap_pair"] for r in rows).items())),
        "direction_distribution": dict(Counter(r["direction"] for r in rows)),
        "multi_label_grounded_rows": sum(r["multi_label_grounded"] == "true" for r in rows),
        "domestic_violence_hidden_fact_rows": sum(r["domestic_violence_hidden_fact"] == "true" for r in rows),
        "paper_failure_mode_rows": sum(bool(r["paper_failure_mode"]) for r in rows),
        "opening_query_words": {
            "min": min(len(r["opening_query"].split()) for r in rows),
            "mean": round(sum(len(r["opening_query"].split()) for r in rows) / len(rows), 2),
            "max": max(len(r["opening_query"].split()) for r in rows),
        },
        "validation_errors": errors,
    }
    if errors:
        raise ValueError("\n".join(errors[:30]))
    return summary


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--legacy", type=Path, default=DEFAULT_LEGACY)
    parser.add_argument("--taxonomy", type=Path, default=DEFAULT_TAXONOMY)
    args = parser.parse_args()

    profile, multi_rows = workbook_profile(args.workbook)
    rows = legacy_rows(args.legacy) + expanded_rows()
    random.Random(20260714).shuffle(rows)
    summary = validate(rows, args.taxonomy)

    write_rows(OUT_CSV, rows)
    OUT_JSONL.parent.mkdir(parents=True, exist_ok=True)
    OUT_JSONL.write_text("".join(json.dumps(row, ensure_ascii=False) + "\n" for row in rows), encoding="utf-8")
    OUT_MULTI.parent.mkdir(parents=True, exist_ok=True)
    with OUT_MULTI.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(multi_rows[0]))
        writer.writeheader(); writer.writerows(multi_rows)
    OUT_PROFILE.write_text(json.dumps({"source_profile": profile, "benchmark_summary": summary}, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(summary, indent=2))


if __name__ == "__main__":
    main()
